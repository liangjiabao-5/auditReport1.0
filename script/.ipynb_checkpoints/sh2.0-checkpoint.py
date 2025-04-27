import pandas as pd
import openpyxl
from openai import OpenAI
from flask import Flask
import os
import time
import sys

app = Flask(__name__)

# ---------------------------------glm4模型--------------------------------------------

base_url = "http://127.0.0.1:8000/v1/"
glm4_client = OpenAI(api_key="EMPTY", base_url=base_url)


def glm4_chat(expectResult, result_record):
    if "\n" in expectResult :
        messages = [
            {
                "role": "system",
                "content": f"""
                按照以下预期结果的几点规则，进行充分的思考和推理，判断user输入的content是否符合:
                请注意：
                1.在评估时对具体对象进行比对，识别和理解它们所属的类别或具有的属性一致则为符合。
                2.预期结果中包含互斥情况。当user输入的content明确支持预期结果中的某一情况时，模型应专注于评估该情况，而无需对其他互斥情况进行判断。
                3.只要语义相同即可，进行模糊匹配。
                
                预期结果：
                {expectResult}

                最后判断，如果每条预期结果都符合则输出"结果：符合"，部分预期结果符合输出"结果：部分符合"，每条预期结果都不符合则输出"结果：不符合"。
                """,
            },
            {
                "role": "user",
                "content": result_record
            }
        ]
    else:
        messages = [
            {
                "role": "system",
                "content": f"""
                按照以下预期结果的规则，判断user输入的content是否符合:
                请注意：
                在评估时对具体对象进行比对，识别和理解它们所属的类别或具有的属性一致则为符合。
                
                预期结果：
                {expectResult}

                最后判断，进行说明，如果预期结果符合content其中的一条内容，则输出"结果：符合"，如果没有任何一条内容符合，则输出"结果：不符合"
                """,
            },
            {
                "role": "user",
                "content": result_record
            }
        ]

#     if "\n" in expectResult :
#         messages = [
#             {
#                 "role": "user",
#                 "content": f"""
#                 作为一个等保测评专家，请按照以下预期结果的几点规则，判断该双引号中的内容是否符合"{result_record}":
#                 请注意：
#                 1.在评估时对具体对象进行比对，识别和理解它们所属的类别或具有的属性一致则为符合。
#                 2.预期结果中包含互斥情况。当user输入的content明确支持预期结果中的某一情况时，模型应专注于评估该情况，而无需对其他互斥情况进行判断。

#                 预期结果：
#                 {expectResult}

#                 最后判断，如果每条预期结果都符合则输出"结果：符合"，部分预期结果符合输出"结果：部分符合"，每条预期结果都不符合则输出"结果：不符合"。
#                 """
#             }
#         ]
#     else:
#         messages = [
#             {
#                 "role": "user",
#                 "content": f"""
#                 作为一个等保测评专家，按照以下预期结果的规则，判断该双引号中的内容是否符合"{result_record}":
#                 请注意：
#                 在评估时对具体对象进行比对，识别和理解它们所属的类别或具有的属性一致则为符合。
                
#                 预期结果：
#                 {expectResult}

#                 最后判断，如果预期结果符合content其中的一条内容，则输出"结果：符合"，如果没有任何一条内容符合，则输出"结果：不符合"
#                 """
#             }
#         ]
            
    use_stream=False
    # print(messages)
    response = glm4_client.chat.completions.create(
        model="glm-4",
        messages=messages,
        stream=use_stream,
        max_tokens=1024,
        # temperature=0.4,
        temperature=0.7,
        presence_penalty=1.2,
        top_p=0.8,
    )
    if response:
        if use_stream:
            for chunk in response:
                print(chunk)
        else:
            return response
    else:
        print("Error:", response.status_code)



#------------------------------------qwen模型---------------------------------------------


openai_api_key = "EMPTY"
openai_api_base = "http://localhost:8000/v1"
qwen_client = OpenAI(
    api_key=openai_api_key,
    base_url=openai_api_base,
)
def qwen_chat(expectResult, result_record):
    
    if "\n" in expectResult :
        messages = [
            {
                "role": "system",
                "content": f"""
                按照以下预期结果的几点规则，判断user输入的content是否符合:
                请注意：
                1.在评估时对具体对象进行比对，识别和理解它们所属的类别或具有的属性一致则为符合。
                2.预期结果中包含互斥情况。当user输入的content明确支持预期结果中的某一情况时，模型应专注于评估该情况，而无需对其他互斥情况进行判断。
                3.只要语义相同即可，进行模糊匹配。
                
                预期结果：
                {expectResult}

                最后判断，如果每条预期结果都符合则输出"结果：符合"，部分预期结果符合输出"结果：部分符合"，每条预期结果都不符合则输出"结果：不符合"。
                """,
            },
            {
                "role": "user",
                "content": result_record
            }
        ]
    else:
        messages = [
            {
                "role": "system",
                "content": f"""
                按照以下预期结果的几点规则，判断user输入的content是否符合:
                请注意：
                在评估时对具体对象进行比对，识别和理解它们所属的类别或具有的属性一致则为符合。
                
                预期结果：
                {expectResult}

                最后判断，进行说明，如果content其中的一条内容符合预期结果，则输出"结果：符合"，如果没有任何一条内容符合，则输出"结果：不符合"。
                """,
            },
            {
                "role": "user",
                "content": result_record
            }
        ]

    chat_response = qwen_client.chat.completions.create(
        model="/mnt/workspace/qwen/model-7B",
        messages=messages,
        temperature=0.7,
        top_p=0.8,
        max_tokens=1024,
        extra_body={
            "repetition_penalty": 1.05,
        }
    )

    return chat_response

# ----------------------------------------------------------------------------------------

def resultDeal(chatResult):
    # print(f"chatResult:{chatResult}")

    # 获取 choices 列表中的第一个 Choice 对象
    first_choice = chatResult.choices[0]

    # 获取 Choice 对象的 message 属性
    message = first_choice.message

    # 获取 message 的 content 字段内容
    content = message.content
    print(f"获取模型执行结果：{content}")

    # # 打印 content 字段的内容
    # if "```python" in content:
    #     result = None
    #     return content, result
    
    # 根据换行符分割字符串
    lines = content.split('\n')
    # 倒序遍历每一行，检查是否包含目标字符
    for line in reversed(lines):
        if "结果：符合" in line:
            result = "符合"
            break;
        elif  "结果：部分符合" in line:
            result = "部分符合"
            break;
        elif "结果：不符合" in line:
            result = "不符合"
            break;
        else:
            result = None
    # print(f"result的结果为：{result}")
    return content, result



sheet = None

def getExpect(securityCategory , controlPoint, quota, index):
    
    global sheet
    # 增加缓存避免每次都取消合并单元格
    if sheet is None:
        # TODO:替换为您想要遍历的目录路径
        directory_path = './workingInstruction'

        # 遍历指定目录及其子目录下的所有文件
        for file in os.listdir(directory_path):
            # print(file)
            # 检查文件名是否以.xlsx结尾，并且是否包含目标字符串
            if file.endswith('.xlsx') and securityCategory in file:
                file_path = os.path.join(directory_path, file)

                # print(f"Reading file: {file_path}")
                break;

        # 打开工作簿
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
         # 取消合并单元格
        unmerge_cells(sheet)


    # 初始化startIndex和endIndex
    startIndex, endIndex = None, None

    # 遍历第二列，寻找controlPoint
    for row in sheet.iter_rows(min_col=2, max_col=2, values_only=False):
        cell = row[0]  # 第二列的单元格
        if cell.value is not None and controlPoint in str(cell.value):  # 检查cell.value是否包含controlPoint
            if startIndex is None:
                startIndex = cell.row  # 记录开始索引
            endIndex = cell.row  # 每次找到时更新结束索引

    if startIndex is not None:
        # print(f"开始索引：{startIndex}, 结束索引：{endIndex}")
        print("-------------------------------------------------------------------")
    else:
        return "未找到控制点"


    # 如果找到了startIndex和endIndex，查找预期结果
    for row in sheet.iter_rows(min_col=1, max_col=5, min_row=startIndex, max_row=endIndex, values_only=False):
        cell2 = row[2]  # 第三列的单元格
        cell5 = row[index - 1]  # 第五列的单元格
        # 检查配额和第五列的值
        if cell2.value is not None and quota in str(cell2.value) and cell5.value is not None:
            # print(cell5.value)
            return cell5.value

# 取消合并单元格
def unmerge_cells(sheet):
    # 只对第二列的单元格进行操作，获取合并单元格范围的列表
    merged_cells = [cell for cell in sheet.merged_cells.ranges if cell.min_col == 2]

    # 遍历合并单元格并取消合并，同时填充内容
    for merged_cell in merged_cells:
        # 获取合并单元格的左上角单元格的值
        value = sheet.cell(row=merged_cell.min_row, column=merged_cell.min_col).value

        # 取消合并单元格
        sheet.unmerge_cells(str(merged_cell))

        # 填充取消合并后的所有单元格
        for row in range(merged_cell.min_row, merged_cell.max_row + 1):
            for col in range(merged_cell.min_col, merged_cell.max_col + 1):
                sheet.cell(row=row, column=col).value = value


if __name__ == "__main__":
    
    llm = sys.argv[1]
    # 假设你的CSV文件名为data.csv，并且第一行是表头
    file_name = sys.argv[2]
    security_class = sys.argv[3]
    # 查看当前工作目录路径
    # current_working_directory = os.getcwd()
    # print("Current working directory:", current_working_directory)

    # 读取CSV文件
    df = pd.read_excel(file_name)

    # 检查'detail'列是否存在于DataFrame中
    if 'detail' in df.columns:
        # 获取'detail'列的所有数据
        detail_data = df['detail']
        #print(detail_data)
    else:
        print("列 'detail' 不存在于数据中。")
    #print(detail_data)

    if 'security_control_class' in df.columns:
        # 获取'detail'列的所有数据
        security_control_class_data = df['security_control_class']
        #print(security_control_class_data)
    else:
        print("列 'security_control_class' 不存在于数据中。")
    #print(security_control_class_data)

    if 'result_record' in df.columns:
        # 获取'detail'列的所有数据
        result_record_data = df['result_record']
        #print(result_record_data)
    else:
        print("列 'result_record' 不存在于数据中。")
    #print(result_record_data)
    
    if 'result_score' in df.columns:
        # 获取'detail'列的所有数据
        result_score_data = df['result_score']
        #print(result_record_data)
    else:
        print("列 'result_score' 不存在于数据中。")


    # 加载已存在的 Excel 文件
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    # 获取现有表格的列名，以确保不会覆盖其他列
    existing_columns = [cell.value for cell in ws[1]]

    # 确保'realy_result'和'reason'列存在，如果没有则添加
    if "realy_result" not in existing_columns:
        ws.cell(row=1, column=len(existing_columns) + 1, value="realy_result")
    if "reason" not in existing_columns:
        ws.cell(row=1, column=len(existing_columns) + 2, value="reason")


    # 获取'realy_result'和'reason'所在列的列号
#     realy_result_col = existing_columns.index("realy_result") + 1
    
    glm4_result_col = existing_columns.index("glm4_result") + 1
    glm4_reason_col = existing_columns.index("glm4_reason") + 1
    
    qwen_result_col = existing_columns.index("qwen_result") + 1
    qwen_reason_col = existing_columns.index("qwen_reason") + 1


    # 初始化一个空的列表来存储结果
    batch_size = 20
    row_start = 2

    # 遍历列表
    for i, data in enumerate(security_control_class_data, start=0):

        # print(result_record_data[i])
        quota = detail_data[i][3:-3]
        expectResult = getExpect(security_class, security_control_class_data[i], quota, 5)
        
        row_num = row_start + i
        
        print(f"第{i + 1}条数据，结果记录为：{result_record_data[i]}------")
        # 调用接口
        if llm == 'qwen':
            qwenChatResult = qwen_chat(expectResult, result_record_data[i])
            qwen_content, qwen_result = resultDeal(qwenChatResult)
            
            if qwen_result != result_score_data[i]:
                operationStep = getExpect(security_class, security_control_class_data[i], quota, 4)
                qwenChatResult = qwen_chat(operationStep, result_record_data[i])
                qwen_content, qwen_result = resultDeal(qwenChatResult)
            
            ws.cell(row=row_num, column=qwen_reason_col, value=qwen_content)
            ws.cell(row=row_num, column=qwen_result_col, value=qwen_result)
        else:
            glm4ChatResult = glm4_chat(expectResult, result_record_data[i])
            glm4_content, glm4_result = resultDeal(glm4ChatResult)
            
            if glm4_result != result_score_data[i]:
                operationStep = getExpect(security_class, security_control_class_data[i], quota, 4)
                glm4ChatResult = glm4_chat(operationStep, result_record_data[i])
                glm4_content, glm4_result = resultDeal(glm4ChatResult)
            
            ws.cell(row=row_num, column=glm4_reason_col, value=glm4_content)
            ws.cell(row=row_num, column=glm4_result_col, value=glm4_result)
        
        
        # TODO: glm4和qwen对比
        # ws.cell(row=row_num, column=realy_result_col, value=result)

        # 每50次调用后保存文件
        if i % batch_size == 0 or i == len(security_control_class_data):
            wb.save(file_name)

    # 保存工作簿
    wb.save(file_name)